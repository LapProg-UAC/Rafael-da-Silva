import random 
from openpyxl import workbook as opyxl

def convert_txt_list(filename:str)->list:
    """
    transforma um ficheiro .txt com nomes de medicamentos numa lista

    parametro filename: o nome do ficheiro .txt a ser convertido
    return: lista com os nomes dos medicamentos
    """
    with open(filename, 'r', encoding='utf-8') as f:
        return f.readlines() 

def generate_intercection_matriz(medicines:list)->list:
    """
    creat a list to generate random reaction between the medicines using numbers 0 to 6

    param medicine: list with all medicines that will react with each other
    return: matriz whit all reactions
    """
    size = len(medicines)
    matriz = [[0]*size for i in range(size)]

    for i in range(size):
        for j in range(i+1, size):
            valor = random.randint(0, 6)
            matriz[i][j] = valor
            matriz[j][i] = valor

    return matriz

def gerar_excel(filename:str, medicines:list, matriz:list)->None:
    """
    generate excel to represente the reactions created in "generate_intercection_matriz"

    param filename: the name of the resulting .xlsx file
    param medicines: same list with name of medicines used in "generate_intercection_matriz"
    param matriz: matriz resulted of the function "generate_intercection_matriz"
    retrun: nothing
    """
    wb = opyxl.Workbook()
    ws = wb.active
    ws.title = "Interações"

    # escrever cabeçalhos das colunas
    for col, med in enumerate(medicines, start=2):
        ws.cell(row=1, column=col, value=med)

    # escrever cabeçalhos das linhas + valores
    for row, med in enumerate(medicines, start=2):
        ws.cell(row=row, column=1, value=med)
        for col in range(len(medicines)):
            ws.cell(row=row, column=col+2, value=matriz[row-2][col])

    wb.save(filename)
    return None
def creat_prescription(medicines:list)->list:
    """
    randomly chose 1 to 10 medicine to creat a prescription

    param medicine: list of medicines
    retrun: list with medicine prescripted
    """
    nummedicine = random.randint(1,10)
    prescription = []
    for i in range(nummedicine):
        prescription.append(random.choice(medicines).strip())
    return prescription


def danger_avaliation(prescription:list,matiz:list)->int:
    """
    check the highs risk reaction between the medicine in the prescription

    param medicine: list of medicines choosed
    param matriz: list with the level of danger of the reaction between de medicines
    retrun: value that represente the level of danger
    """
    danger = []
    max = 0
    for i in range(len(prescription)):
        for j in range(i + 1,len(prescription)):
            danger.append(matriz[i][j])
    for i in range(len(danger)):
        if max < danger[i]:
            max = danger[i]
    return max 


medicines = convert_txt_list("medicines.txt")
matriz = generate_intercection_matriz(medicines)
gerar_excel("interacoes.xlsx", medicines, matriz)
prescription = creat_prescription(medicines)
print("a prescriçao dada foi:", *prescription,sep="\n")
print(f"o nivel de perigo dessa presgriçao é {danger_avaliation(prescription,matriz)}")