import random
import openpyxl
import json
import encriptacao

CHAVE = [3, 5, 2, 9]

def AbrirFicheiroMedicamentos(Caminho: str):
    try:
        with open(Caminho, "r", encoding="utf-8") as f:
            conteudo = f.read().replace("", "")
            return [l.strip() for l in conteudo.splitlines() if l.strip()]
    except FileNotFoundError:
        print("Erro: O ficheiro", Caminho, "não encontrado.")
        return []

def CriarTabelaSinóptica(Medicamentos: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabela Sinóptica"

    RegrasInteracao = {}

    for i, Nome_i in enumerate(Medicamentos, start=2):
        ws.cell(row=i, column=1, value=Nome_i)
        ws.cell(row=1, column=i, value=Nome_i)

        for j, Nome_j in enumerate(Medicamentos, start=2):

            if i == j:
                Valor = 0
            else:
                Valor = random.randint(0, 6)

            ws.cell(row=i, column=j, value=Valor)
            RegrasInteracao[(Nome_i, Nome_j)] = Valor

    return wb, RegrasInteracao

def SalvarTabelaSinóptica(NomeFicheiro: str, wb: openpyxl.Workbook):
    try:
        wb.save(NomeFicheiro + ".xlsx")

        with open(NomeFicheiro + ".txt", "w", encoding="utf-8") as f:
            for row in wb.active.iter_rows(values_only=True):
                f.write("\t".join(str(cell) if cell is not None else "" for cell in row) + "\n")

        print(f"Tabela salva como {NomeFicheiro}.xlsx e .txt")

    except Exception as e:
        print("Erro ao salvar:", e)

def GerarDadosFinais(CaminhoNomes: str, CaminhoSobrenomes: str, NumUtentes: int, ListaMedicamentos: list, Regras: dict, CHAVE):
    NomesBase = AbrirFicheiroMedicamentos(CaminhoNomes)
    SobrenomesBase = AbrirFicheiroMedicamentos(CaminhoSobrenomes)

    ListaIDs = []
    ListaJson = []

    # Gerar IDs únicos
    i = 0
    while i < NumUtentes:
        NumeroID = random.randint(1000, 9999)
        if NumeroID not in ListaIDs:
            ListaIDs.append(NumeroID)
            i += 1

    for uid in ListaIDs:
        NomeCompleto = f"{random.choice(NomesBase)} {random.choice(SobrenomesBase)}"
        Receita = random.sample(ListaMedicamentos, k=random.randint(3, 5))

        SomaInteracoes = 0
        Interage = False

        for i in range(len(Receita)):
            for j in range(i + 1, len(Receita)):
                Grau = Regras.get((Receita[i], Receita[j]), 0)
                SomaInteracoes += Grau
                if Grau > 0:
                    Interage = True

        # Usar a função encriptar do módulo encriptação
        id_encriptado = encriptacao.encriptar(str(uid), CHAVE)

        ListaJson.append({
            "id_utente": id_encriptado,
            "nome_utente": NomeCompleto,
            "receita": Receita,
            "balanco_total": SomaInteracoes,
            "resultado": "Com Interações" if Interage else "Sem Interações"
        })

    return ListaJson

def SalvarJSON(Dados: list, Nome: str):

    with open(Nome, "w", encoding="utf-8") as f:
        json.dump(Dados, f, indent=4, ensure_ascii=False)

    print(f"Ficheiro JSON '{Nome}' criado.")

def main():
    CaminhoMeds = "medicamentos.txt"
    CaminhoNomes = "nomes.txt"
    CaminhoSobr = "apelidos.txt"

    Medicamentos = AbrirFicheiroMedicamentos(CaminhoMeds)

    workbook, RegrasMemoria = CriarTabelaSinóptica(Medicamentos)

    SalvarTabelaSinóptica("ResultadoSinoptico", workbook)

    DadosFinais = GerarDadosFinais(CaminhoNomes, CaminhoSobr, 10, Medicamentos, RegrasMemoria)

    SalvarJSON(DadosFinais, "FicheiroFinal.json")

if __name__ == "__main__":
    main()